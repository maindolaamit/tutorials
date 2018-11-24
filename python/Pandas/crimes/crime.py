import pandas as pd 
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Use the templatae file to create a Crime chart for a region

wb = load_workbook('template.xlsx')
ws = wb.active

# Give the data type explicitly to save the time 
df = pd.read_csv('crime.csv', encoding='utf-8', dtype={"INCIDENT_NUMBER":str, "OFFENSE_CODE": str, "OFFENSE_CODE_GROUP": str
                , "OFFENSE_DESCRIPTION":str, "DISTRICT":str, "REPORTING_AREA":str, "SHOOTING":str, "YEAR":str, "MONTH":str
                , "DAY_OF_WEEK":str, "HOUR":str})
        
# Search for a crime
# ws['G6'].value = "No. of Vandalism Crimes by District and Year"
df1 = df[df["OFFENSE_CODE_GROUP"] == "Vandalism"]

# Replace null value
df1 = df1.replace(np.nan, 'NA', regex=True)

# Get the stats
total_crimes = len(df.index)
vandalism_count = len(df1.index)
vandalism_percent = (vandalism_count / total_crimes) * 100

# Save the value
ws['O8'].value = total_crimes
ws['P8'].value = vandalism_count
ws['Q8'].value = round(vandalism_percent, 2)

df2 = df1.groupby(['DISTRICT','YEAR']).count()['INCIDENT_NUMBER'].unstack(level=0)
print(df2.head())

rows = dataframe_to_rows(df2)
for rid, row in enumerate(rows, 10):
    for cid, col in enumerate(row, 1):
        ws.cell(row=rid, column=cid, value=col)


wb.save('Results.xlsx')