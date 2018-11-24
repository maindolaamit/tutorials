import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load excel files in the DataFrame
df = pd.read_excel('total_sales.xlsx')

print(df.head())

wb = load_workbook('regions.xlsx')
ws = wb.active


rows = dataframe_to_rows(df[['Region', 'Sales Rep', 'Total']], index=False)
for rid, row in enumerate(rows, 1):
    for cid, col in enumerate(row, 7):
        ws.cell(row=rid, column=cid, value=col)

wb.save('Result.xlsx')
print('Data written successfully !')
