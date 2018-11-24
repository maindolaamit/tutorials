import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Load excel files in the DataFrame
df_1 = pd.read_excel('sales.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('sales.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('sales_3.xlsx')

# Combine the DataFrames
df = pd.concat([df_1, df_2, df_3], sort=False, ignore_index=True)

# View the values in dataframe at location 50, If we do not ignore the index then we will get 3 values from 3 DataFrames
# print(df.loc[50])

# Let's view group the data by Item
print(df.groupby(['Item']).sum()['Units Sold'])

# Create a new Total Column 
df['Total'] = df['Cost per'] * df['Units Sold']

# Save the data into a new Excel file 
df.to_excel("total_sales.xlsx", index=None)

# Now let's open the Workbook for some formatting
wb = load_workbook('total_sales.xlsx')
ws = wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)

wb.save()